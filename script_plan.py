import csv
import math
from argparse import ArgumentParser
from datetime import datetime, timedelta
from os import getcwd
from os.path import join
from xml.etree.ElementTree import parse

import yaml
from openpyxl import load_workbook
from tqdm import tqdm

from logic.xml_tools import get_text_value, get_float_value_with_dot
from tools.ia_rest import IAImportExport
from tools.listofdicts_to_csv import dict2csv

# WIP_DEPT_TO_SKIP = ['1011', '1015', '1012', '1013',
#                     '1014', '1019', '1017', '0716', 'N101', '7403']
# DEPT_TO_SKIP = ['1011', '1015', '1012', '1013',
#                 '1014', '1019', '1017', '0716', 'N101', '7403']
# WIP_DEPT_TO_SKIP = []
DEPT_TO_SKIP = []


def read_plan_from_file(xml_file, skip):
    tree = parse(xml_file)
    root = tree.getroot()
    report = {}
    WIP_DEPT_TO_SKIP = skip
    # читаем все фреймы в исходном файле
    rows = root.findall('MAT_DATA')

    for material in tqdm(rows, desc='Считываем XML'):
        # читаем идентификатор ДСЕ
        product = get_text_value(material, "MATNR")
        if '-' in product:
            product = f'{product}_Z{product[-5]}01'
        report[product] = {
            'PLAN': {},
            'WIP': {}
        }
        # читаем все строки НЗП по этой ДСЕ
        wip_rows = material.findall('MT_STOCK')
        # читаем все заказы плана по этой ДСЕ
        plan_rows = material.findall('MT_CHAIN')
        for stock in wip_rows:
            # if str(get_text_value(stock, 'LGORT'))[:3] != '105' and '-' not in product:
            #     skip = True
            if '178033017001' in product:
                a = 1
            skip = False
            # if get_text_value(stock, 'LGORT') is None:
            #     a = 1
            # игнорирую НЗП в некоторых подразделениях (это НЗП сборки)
            for dept in WIP_DEPT_TO_SKIP:
                if dept in str(get_text_value(stock, 'LGORT')):
                    skip = True
                    break
            if str(get_text_value(stock, 'LGORT')) == '1016' and '-' not in product:
                skip = True
            if skip:
                continue
            # придумаем как назвать идентификатор партии
            # условие -- ID должно быть уникальным.
            # В нашем случае ДСЕ-Подразделение где лежит НЗП обещает
            # быть уникально в виду логики формирования этого НЗП
            batch_id = '{}_{}'.format(
                product,
                get_text_value(stock, 'LGORT')
            )
            # В LABST лежит количество
            amount = get_float_value_with_dot(stock, 'LABST')
            report[product]['WIP'][batch_id] = amount
        for order in plan_rows:
            skip = False
            # В план попали заказы, которые делаем не мы --
            # их убираем по паттерну в названии. Могу попросить просто
            # не выгружать
            for dept in DEPT_TO_SKIP:
                if dept in get_text_value(order, 'TO').replace('.', ''):
                    skip = True
                    break
            # if '101' in get_text_value(order, 'FROM'):
            #     if '101.6' not in get_text_value(order, 'FROM'):
            #         skip = True
            if skip:
                continue
            # название заказа сформировали за нас
            order_name = get_text_value(order, 'ORDER')
            # записываем количество из поля AMOUNT -- но оно
            # посчитано ими неправильно, потом будет логика
            # обработки этих данных еще
            report[product]['PLAN'][order_name] = {
                'DATE': get_text_value(order, 'DATE_TO'),
                'AMOUNT': get_float_value_with_dot(order, 'AMOUNT')
            }

    return report


def read_variance(path, sheet):

    print('Читаем файл {} с отклонениями НЗП'.format(path))

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    header = [cell.value for cell in ws[1]]
    report = []

    for row in ws.iter_rows(min_row=2):
        values = {}
        for key, cell in zip(header, row):
            values[key] = cell.value
        report.append(values)

    result = {}

    for row in report:
        if str(row['Участок']).zfill(4) in DEPT_TO_SKIP:
            continue
        if row['Отклонение'] is None:
            continue
        result['{}_{}'.format(
            str(row['Номенклатурный номер']).zfill(18),
            row['Участок']
        )] = round(row['Отклонение']*1000)/1000

    return result


def main():
    from logic import read_from_ftp

    parser = ArgumentParser(
        description='Инструмент консольной генерации отчетов '
                    'по результатам моделирования.'
    )

    parser.add_argument('-c', '--config', required=False,
                        default=join(getcwd(), 'kk_plan_parser.yml'))
    parser.add_argument('-s', '--server', required=False,
                        default=join(getcwd(), 'server.yml'))
    args = parser.parse_args()

    with open(args.config, 'r', encoding="utf-8") as stream:
        config = yaml.load(stream, Loader=yaml.SafeLoader)

    with open(args.server, 'r', encoding="utf-8") as stream:
        server = yaml.load(stream, Loader=yaml.SafeLoader)

    with IAImportExport.from_config(config['instance']) as session:
        report = []
        special_items = []
        all_items = set()
        for row in session.get_from_rest_collection('entity'):
            if row['entity_type_id'] is None:
                continue
            all_items.add(row['identity'])
            if '-' in row['identity'] and '(' not in row['identity']:
                report.append({
                    'OP_ID': row['identity'],
                    'CODE': row['identity']
                })
                special_items.append(row['identity'])
        dict2csv(report, 'special_items.csv')

    sftpURL = server['sftpURL']
    sftpUser = server['sftpUser']
    sftpPass = server['sftpPass']
    # sftpUser = 'a.a.stolov'
    # sftpPass = 'Yunku_Kk2021kK'
    # sftpPath = '/home/http_request_collector/app/data/input/state/POST'
    sftpPath = config['path']
    wipskip = config['skip']

    xml_data = read_from_ftp.read_plan_from_ftp(
        sftpURL, sftpUser, sftpPass, sftpPath, wipskip, asm=False
    )

    plan = []
    wip = []

    for entity in tqdm(xml_data, desc='Разбор данных'):
        # Возможны два варианта MATNR у ДСЕ:
        # 1. 18 символов (преимущественно цифр) с 6 нулями впереди --
        # это прямо ДСЕ
        # 2. 12 символов, дефис, 5 символов -- это полуфабрикаты той ДСЕ,
        # которая в первых 12 символах
        #
        # текстовой логикой снизу я из кода полуфабриката получаю код ДСЕ
        if '-' in entity:
            product = entity[:12].zfill(18)
        else:
            product = entity

        for item in special_items:
            if item in entity:
                product = entity[:18]
                break

        acc_amount = None
        for num, order in enumerate(xml_data[entity]['PLAN']):
            if acc_amount is None:
                # первое значение поля amount назначаем, как
                # количество "накопленное"
                acc_amount = xml_data[entity]['PLAN'][order]['AMOUNT']
            elif acc_amount < 0:
                # если накопленное количество с предыдущих заказов меньше
                # нуля, то у нас есть профицит, прибавляем к нему потребность
                # на эту неделю
                acc_amount += xml_data[entity]['PLAN'][order]['AMOUNT']
            else:
                # а если больше нуля, то просто присваиваем ему значение
                # потребности текущей недели -- значит профицита с прошлой
                # недели нет
                if num == 1 and xml_data[entity]['PLAN'][order]['AMOUNT'] < 0:
                    plan[-1]['AMOUNT'] += xml_data[entity]['PLAN'][order]['AMOUNT']
                    acc_amount = min(0, plan[-1]['AMOUNT'])
                else:
                    acc_amount = xml_data[entity]['PLAN'][order]['AMOUNT']
            # логику можно упростить:
            # - начинаем накопленное количество с нуля по позиции
            # - прибавляем к накопленному количеству AMOUNT
            # - если накопленное количество больше 0, то добавляем
            # в результирующий план заказ с накопленным количеством, а
            # накопленное количество обнуляем
            # - переходим на следующую строку и переходим к п.2
            date = datetime.strptime(
                xml_data[entity]['PLAN'][order]['DATE'],
                '%Y-%m-%d %H:%M:%S'
            )
            date_from = max(
                date - timedelta(days=21),
                datetime.now().replace(hour=7, minute=0,
                                       second=0, microsecond=0)
            )
            plan.append({
                # в заказе режу незначащие пробелы в начале
                'ORDER': order.strip(),
                'CODE': product,
                'DATE_FROM': date_from,
                'DATE_TO': date,
                # поле init_amount -- для дебага
                'INIT_AMOUNT': xml_data[entity]['PLAN'][order]['AMOUNT'],
                # теоретически могут быть дробные количества (причин не знаю),
                # округляю до большего целого
                'AMOUNT': math.ceil(acc_amount)
            })
        # if '107201050001' in product:
            # print(entity)
            # print(product)
            # print(entity in all_items)
            # print(product in all_items)
            # a = 1
        # if product not in all_items:
        #     continue
        for batch in xml_data[entity]['WIP']:
            if entity[:18] != product:
                wip.append({
                    'BATCH_ID': batch,
                    'CODE': product,
                    'INIT_AMOUNT': xml_data[entity]['WIP'][batch],
                    'AMOUNT': round(xml_data[entity]['WIP'][batch] * 1000)/1000,
                    'OPERATION_ID': entity,
                    'OPERATION_PROGRESS': 0,
                    'ORDER': ''
                })
            else:
                wip.append({
                    'BATCH_ID': batch,
                    'CODE': product,
                    'INIT_AMOUNT': xml_data[entity]['WIP'][batch],
                    'AMOUNT': round(
                        (xml_data[entity]['WIP'][batch])*1000
                    )/1000,
                    'OPERATION_ID': '',
                    'OPERATION_PROGRESS': 100,
                    'ORDER': ''
                })

    # сейчас нет у меня никакой сложной логики, чтоб определить текущие
    # приоритеты заказов -- по-умолчанию сортирую сначала по дате
    # (по возрастанию), потом по количеству (по убыванию)
    try:
        plan = sorted(plan, key=lambda x: (x['DATE_TO'], -x['AMOUNT']))
        keys = plan[0].keys()
        with open(config['output']['plan'], 'w', newline='', encoding='utf-8') as output_file:
            dict_writer = csv.DictWriter(output_file, keys)
            dict_writer.writeheader()
            dict_writer.writerows(plan)
    except IndexError:
        tqdm.write('План не создан')

    keys = wip[0].keys()
    with open(config['output']['wip'], 'w', newline='', encoding='utf-8') as output_file:
        dict_writer = csv.DictWriter(output_file, keys)
        dict_writer.writeheader()
        dict_writer.writerows(wip)


if __name__ == '__main__':
    main()
